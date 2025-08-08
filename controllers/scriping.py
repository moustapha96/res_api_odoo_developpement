# from odoo import http
# from odoo.http import request, Response
# import json
# import requests
# from bs4 import BeautifulSoup
# import werkzeug.wrappers

# import os
# import datetime
# import urllib.parse
# from pathlib import Path

# import base64
# import logging
# import io
# from PIL import Image

# from openpyxl import load_workbook

# # import pandas as pd


# _logger = logging.getLogger(__name__)


# class ScraperController(http.Controller):

#     @http.route('/api/scrape_products_full', methods=['GET'], type='http', auth='none', cors="*")
#     def scrape_products_full(self, **kwargs):
#         target_url = kwargs.get('target_url')
#         fetch_details = kwargs.get('fetch_details', 'false').lower() == 'true'
#         save_to_file = kwargs.get('save_to_file', 'true').lower() == 'true'

#         if not target_url:
#             return werkzeug.wrappers.Response(
#                 status=400,
#                 content_type='application/json; charset=utf-8',
#                 headers=[('Cache-Control', 'no-store'), ('Pragma', 'no-cache')],
#                 response=json.dumps({"error": "Veuillez fournir l'URL cible via le paramètre 'target_url'."})
#             )

#         try:
#             response = requests.get(target_url, timeout=10)
#             response.raise_for_status()
#             soup = BeautifulSoup(response.text, 'html.parser')
#             products = soup.select("ul.products li.product")
#             product_data = []

#             for product in products:
#                 title = product.select_one("h2.woocommerce-loop-product__title")
#                 link = product.select_one("a.woocommerce-LoopProduct-link")
#                 image = product.select_one("img")
#                 title_text = title.text.replace('  ', ' ') 

#                 product_info = {
#                     "title": title_text.strip() if title else None,
#                     "link": link["href"] if link else None,
#                     "image" : image.get("src") if image else None,
#                 }


#                 # Détails du produit
#                 if fetch_details and product_info["link"]:
#                     try:
#                         details_produits = self.fetch_page_text(target_url=product_info["link"])
#                         # product_info["details"] = details_produits
#                         product_info["summary_text"] = details_produits.get("summary_text", None)
#                         product_info["images"] = details_produits.get("images", [])
#                         product_info["categories"] = details_produits.get("categories", [])
#                         product_info["tags"] = details_produits.get("tags", [])

#                     except Exception as e:
#                         product_info["detail_fetch_error"] = str(e)

#                 product_data.append(product_info)

#             page_data = {
#                 "url": target_url,
#                 "scrape_date": datetime.datetime.now().isoformat(),
#                 "produits_nbre": len(product_data),
#                 "produits": product_data
#             }

#             product_titles = [p['title'] for p in product_data]
#             existing_products = request.env['product.product'].sudo().search([('name', 'in', product_titles)])
#             existing_product_titles = set(existing_products.mapped('name'))

#             for product in product_data:
#                 product['exists_in_db'] = product['title'] in existing_product_titles

#             file_path = None
#             if save_to_file:
#                 data_dir = os.path.join(os.path.dirname(__file__), '../data')
#                 os.makedirs(data_dir, exist_ok=True)

#                 domain = urllib.parse.urlparse(target_url).netloc.replace('.', '_')

#                 parsed_url = urllib.parse.urlparse(target_url)
#                 category = parsed_url.path.split('/')[-2] if parsed_url.path.count('/') > 1 else 'uncategorized'
             
#                 domain = parsed_url.netloc.replace('.', '_')
#                 timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
#                 filename = f"products_{domain}_{category}.json"
#                 # filename = f"products_{domain}_{category}_{timestamp}.json"
#                 file_path = os.path.join(data_dir, filename)

#                 with open(file_path, 'w', encoding='utf-8') as f:
#                     json.dump(page_data, f, ensure_ascii=False, indent=4)

#                 page_data["file_saved"] = file_path

#             return werkzeug.wrappers.Response(
#                 status=200,
#                 content_type='application/json; charset=utf-8',
#                 headers=[('Cache-Control', 'no-store'), ('Pragma', 'no-cache')],
#                 response=json.dumps(page_data, ensure_ascii=False, indent=4)
#             )

#         except requests.exceptions.RequestException as e:
#             return werkzeug.wrappers.Response(
#                 status=500,
#                 content_type='application/json; charset=utf-8',
#                 headers=[('Cache-Control', 'no-store'), ('Pragma', 'no-cache')],
#                 response=json.dumps({"error": f"Erreur lors de la requête : {str(e)}"})
#             )
#         except Exception as e:
#             return werkzeug.wrappers.Response(
#                 status=500,
#                 content_type='application/json; charset=utf-8',
#                 headers=[('Cache-Control', 'no-store'), ('Pragma', 'no-cache')],
#                 response=json.dumps({"error": f"Erreur inconnue : {str(e)}"})
#             )


#     def fetch_page_text(self,target_url):
#         """Fetches the text from a specific section of a webpage."""
#         if not target_url:
#             return {
#                 "error": "Veuillez fournir l'URL cible via le paramètre 'target_url'."
#             }

#         try:
#             response = requests.get(target_url, timeout=10)
#             response.raise_for_status()
#             soup = BeautifulSoup(response.text, 'html.parser')

#             # Extraire le texte de la section spécifique
#             summary_section = soup.select_one('div.summary.entry-summary')
#             if not summary_section:
#                 return {
#                     "error": "La section spécifiée n'a pas été trouvée."
#                 }

#             # Extraire le titre
#             title_element = summary_section.select_one('h1.product_title.entry-title')
#             title_text = title_element.get_text(strip=True) if title_element else "Titre non trouvé"

#             # Extraire la description courte
#             description_section = summary_section.select_one('div.woocommerce-product-details__short-description')
#             if description_section:
#                 paragraphs = description_section.find_all('p')
#                 summary_text = "\n".join(p.get_text(strip=True) for p in paragraphs)
#             else:
#                 summary_text = summary_section.get_text(separator='\n', strip=True)

#             # Extraire les images de la galerie
#             gallery_section = soup.select_one('div.woocommerce-product-gallery')
#             images = []
#             if gallery_section:
#                 for img in gallery_section.select('div.woocommerce-product-gallery__image img'):
#                     images.append({
#                         "src": img.get("data-large_image") or img.get("src")
#                     })

#             # Extraire la catégorie et les tags
#             category_element = summary_section.select_one('span.posted_in')
#             categories = [a.get_text(strip=True) for a in category_element.select('a')] if category_element else []

#             tag_element = summary_section.select_one('span.tagged_as')
#             tags = [a.get_text(strip=True) for a in tag_element.select('a')] if tag_element else []

#             page_data = {
#                 "title": title_text,
#                 "summary_text": summary_text,
#                 "images": images,
#                 "categories": categories,
#                 "tags": tags
#             }

#             return page_data

#         except requests.exceptions.RequestException as e:
#             return {
#                 "error": f"Erreur lors de la requête : {str(e)}"
#             }
#         except Exception as e:
#             return {
#                 "error": f"Erreur inconnue : {str(e)}"
#             }
        
#     @http.route('/api/read_product_files', methods=['GET'], type='http', auth='none', cors="*")
#     def read_product_files(self, **kwargs):
#         file_name = kwargs.get('file_name')

#         if not file_name:
#             return werkzeug.wrappers.Response(
#                 status=400,
#                 content_type='application/json; charset=utf-8',
#                 headers=[('Cache-Control', 'no-store'), ('Pragma', 'no-cache')],
#                 response=json.dumps({"error": "Veuillez fournir le nom du fichier via le paramètre 'file_name'."})
#             )

#         data_dir = os.path.join(os.path.dirname(__file__), '../data')
#         file_path = os.path.join(data_dir, file_name)

#         if not os.path.exists(file_path):
#             return werkzeug.wrappers.Response(
#                 status=404,
#                 content_type='application/json; charset=utf-8',
#                 headers=[('Cache-Control', 'no-store'), ('Pragma', 'no-cache')],
#                 response=json.dumps({"error": "Fichier non trouvé."})
#             )

#         try:
#             with open(file_path, 'r', encoding='utf-8') as f:
#                 file_data = json.load(f)

#             return werkzeug.wrappers.Response(
#                 status=200,
#                 content_type='application/json; charset=utf-8',
#                 headers=[('Cache-Control', 'no-store'), ('Pragma', 'no-cache')],
#                 response=json.dumps(file_data['produits'], ensure_ascii=False, indent=4)
#             )

#         except Exception as e:
#             return werkzeug.wrappers.Response(
#                 status=500,
#                 content_type='application/json; charset=utf-8',
#                 headers=[('Cache-Control', 'no-store'), ('Pragma', 'no-cache')],
#                 response=json.dumps({"error": f"Erreur lors de la lecture du fichier : {str(e)}"})
#             )


#     @http.route('/api/read_product_file', methods=['GET'], type='http', auth='none', cors="*")
#     def read_product_file(self, **kwargs):
#         file_name = kwargs.get('file_name')

#         if not file_name:
#             return werkzeug.wrappers.Response(
#                 status=400,
#                 content_type='application/json; charset=utf-8',
#                 headers=[('Cache-Control', 'no-store'), ('Pragma', 'no-cache')],
#                 response=json.dumps({"error": "Veuillez fournir le nom du fichier via le paramètre 'file_name'."})
#             )

#         data_dir = os.path.join(os.path.dirname(__file__), '../data')
#         file_path = os.path.join(data_dir, file_name)

#         if not os.path.exists(file_path):
#             return werkzeug.wrappers.Response(
#                 status=404,
#                 content_type='application/json; charset=utf-8',
#                 headers=[('Cache-Control', 'no-store'), ('Pragma', 'no-cache')],
#                 response=json.dumps({"error": "Fichier non trouvé."})
#             )

#         try:
#             with open(file_path, 'r', encoding='utf-8') as f:
#                 file_data = json.load(f)

#             produits = file_data.get('produits', [])
#             created_products = []
#             updated_products = []
#             skipped_products = []
            
#             # Obtenir la catégorie par défaut avant de commencer
#             default_category_id = self.ensure_default_category()
#             _logger.info(f"Catégorie par défaut ID: {default_category_id}")
            
#             # Utiliser un savepoint pour pouvoir annuler en cas d'erreur
#             with request.env.cr.savepoint():
#                 for produit in produits:
#                     if produit.get('title'):
#                         _logger.info(f"Traitement du produit: {produit['title']}")
                        
#                         try:
#                             # Vérifier si le modèle de produit existe déjà
#                             existing_product_template = request.env['product.template'].sudo().search(
#                                 [('name', '=', produit['title'])], limit=1)
                            
#                             if existing_product_template:
#                                 _logger.info(f"Modèle de produit existant trouvé: {existing_product_template.id}")
#                                 # Le modèle de produit existe déjà, nous n'avons pas besoin de créer un nouveau produit
#                                 updated_products.append(produit['title'])
#                             else:
#                                 _logger.info(f"Création d'un nouveau modèle de produit: {produit['title']}")
                                
#                                 # Vérifier et récupérer l'image
#                                 image_base64 = self.get_image_base64(produit.get('image'))
                                
#                                 # Récupérer les images supplémentaires
#                                 images = produit.get('images', [])
#                                 image_base64_list = []
                                
#                                 # Traiter chaque image avec gestion des erreurs
#                                 for img in images:
#                                     if img.get('src'):
#                                         img_base64 = self.get_image_base64(img['src'])
#                                         if img_base64:
#                                             image_base64_list.append(img_base64)
                                
#                                 # S'assurer que categ_id est toujours défini
#                                 categ_id = self.get_category_id(produit.get('categories', []))
#                                 if not categ_id:
#                                     categ_id = default_category_id
                                
#                                 _logger.info(f"Catégorie utilisée pour le produit: {categ_id}")
                                
#                                 # Création d'un nouveau modèle de produit
#                                 product_vals = {
#                                     'name': produit['title'],
#                                     'sale_ok': False,
#                                     'purchase_ok': False,
#                                     'detailed_type': 'product',
#                                     'description': produit.get('summary_text'),
#                                     'list_price': float(produit.get('price', 0.0)),
#                                     'standard_price': float(produit.get('cost', 0.0)),
#                                     'en_promo': bool(produit.get('en_promo', False)),
#                                     'is_preorder': bool(produit.get('is_preorder', False)),
#                                     'preorder_price': float(produit.get('preorder_price', 0.0)),
#                                     'promo_price': float(produit.get('promo_price', 0.0)),
#                                     'is_creditorder': bool(produit.get('is_creditorder', False)),
#                                     'creditorder_price': float(produit.get('creditorder_price', 0.0)),
#                                     'rang': int(produit.get('rang', 0)),
#                                     'categ_id': categ_id,  # Utiliser la catégorie récupérée ou la catégorie par défaut
#                                 }
                                
#                                 # Ajouter les tags si disponibles
#                                 tag_ids = self.get_tag_ids(produit.get('tags', []))
#                                 if tag_ids:
#                                     product_vals['product_tag_ids'] = [(6, 0, tag_ids)]
                                
#                                 # Ajouter les images supplémentaires si disponibles
#                                 if image_base64_list:
#                                     if len(image_base64_list) > 0:
#                                         product_vals['image_1'] = image_base64_list[0]
#                                     if len(image_base64_list) > 1:
#                                         product_vals['image_2'] = image_base64_list[1]
#                                     if len(image_base64_list) > 2:
#                                         product_vals['image_3'] = image_base64_list[2]
#                                     if len(image_base64_list) > 3:
#                                         product_vals['image_4'] = image_base64_list[3]

#                                 # N'ajouter l'image principale que si elle est valide
#                                 if image_base64:
#                                     product_vals['image_1920'] = image_base64
                                
#                                 _logger.info(f"Création du produit avec les valeurs: {product_vals.keys()}")
#                                 _logger.info(f"Catégorie ID utilisée: {product_vals.get('categ_id')}")
                                
#                                 product_template = request.env['product.template'].sudo().create(product_vals)
#                                 _logger.info(f"Nouveau modèle de produit créé avec ID: {product_template.id}")
#                                 created_products.append(produit['title'])
#                         except Exception as e:
#                             _logger.error(f"Erreur lors du traitement du produit {produit['title']}: {str(e)}", exc_info=True)
#                             skipped_products.append(f"{produit['title']} (Erreur: {str(e)})")

#             result = {
#                 "success": True,
#                 "created": len(created_products),
#                 "updated": len(updated_products),
#                 "skipped": len(skipped_products),
#                 "created_products": created_products,
#                 "updated_products": updated_products,
#                 "skipped_products": skipped_products
#             }

#             return werkzeug.wrappers.Response(
#                 status=200,
#                 content_type='application/json; charset=utf-8',
#                 headers=[('Cache-Control', 'no-store'), ('Pragma', 'no-cache')],
#                 response=json.dumps(result, ensure_ascii=False, indent=4)
#             )

#         except Exception as e:
#             _logger.error(f"Erreur lors du traitement du fichier: {str(e)}", exc_info=True)
#             return werkzeug.wrappers.Response(
#                 status=500,
#                 content_type='application/json; charset=utf-8',
#                 headers=[('Cache-Control', 'no-store'), ('Pragma', 'no-cache')],
#                 response=json.dumps({"error": f"Erreur lors de la lecture du fichier : {str(e)}"})
#             )


#     def ensure_default_category(self):
#         """
#         S'assure qu'une catégorie par défaut existe et retourne son ID.
#         Cette méthode ne devrait jamais retourner False ou None.
#         """
#         # Catégorie par défaut à utiliser
#         default_category_name = "Non catégorisé"
        
#         # Rechercher la catégorie par défaut
#         default_category = request.env['product.category'].sudo().search([('name', '=', default_category_name)], limit=1)
        
#         if not default_category:
#             # Créer la catégorie par défaut si elle n'existe pas
#             _logger.info(f"Création de la catégorie par défaut: {default_category_name}")
#             try:
#                 default_category = request.env['product.category'].sudo().create({'name': default_category_name})
#                 _logger.info(f"Catégorie par défaut créée avec ID: {default_category.id}")
#             except Exception as e:
#                 _logger.error(f"Erreur lors de la création de la catégorie par défaut: {str(e)}")
#                 # En cas d'échec, utiliser la catégorie "All" (ID 1) qui existe toujours dans Odoo
#                 return 1
        
#         _logger.info(f"Catégorie par défaut trouvée avec ID: {default_category.id}")
#         return default_category.id


#     def get_category_id(self, categories):
#         """
#         Récupère ou crée une catégorie de produit.
#         Retourne un ID de catégorie valide ou None si une erreur se produit.
#         """
#         try:
#             if not categories or not categories[0]:
#                 _logger.warning("Aucune catégorie fournie")
#                 return None
            
#             category_name = categories[0]  # Prendre la première catégorie
#             _logger.info(f"Recherche de la catégorie: {category_name}")
            
#             category = request.env['product.category'].sudo().search([('name', '=', category_name)], limit=1)
            
#             if not category:
#                 # Créer la catégorie si elle n'existe pas
#                 _logger.info(f"Création de la catégorie: {category_name}")
#                 category = request.env['product.category'].sudo().create({'name': category_name})
            
#             _logger.info(f"Catégorie trouvée/créée avec ID: {category.id}")
#             return category.id
#         except Exception as e:
#             _logger.error(f"Erreur lors de la récupération/création de la catégorie: {str(e)}")
#             return None



#     def get_tag_ids(self, tags):
#         """
#         Récupère ou crée des tags de produit.
#         Retourne une liste d'IDs de tags.
#         """
#         tag_ids = []
#         if not tags:
#             return tag_ids
            
#         for tag_name in tags:
#             try:
#                 tag = request.env['product.tag'].sudo().search([('name', '=', tag_name)], limit=1)
#                 if tag:
#                     tag_ids.append(tag.id)
#                 else:
#                     new_tag = request.env['product.tag'].sudo().create({'name': tag_name})
#                     tag_ids.append(new_tag.id)
#             except Exception as e:
#                 _logger.error(f"Erreur lors de la récupération/création du tag {tag_name}: {str(e)}")
                
#         return tag_ids


#     def get_image_base64(self, image_url):
#         """Récupère une image depuis une URL et la convertit en base64 avec vérification."""
#         if not image_url:
#             _logger.warning("Aucune URL d'image fournie")
#             return False
            
#         try:
#             _logger.info(f"Téléchargement de l'image depuis: {image_url}")
#             response = requests.get(image_url, timeout=10)
#             response.raise_for_status()
            
#             # Vérifier si le contenu est une image valide
#             content = response.content
#             try:
#                 # Essayer d'ouvrir l'image avec PIL pour vérifier qu'elle est valide
#                 img = Image.open(io.BytesIO(content))
#                 img.verify()  # Vérifier que l'image est valide
                
#                 # Réouvrir l'image pour la traiter (après verify(), l'image est fermée)
#                 img = Image.open(io.BytesIO(content))
                
#                 # Convertir en RGB si nécessaire (pour les images avec canal alpha)
#                 if img.mode in ('RGBA', 'LA') or (img.mode == 'P' and 'transparency' in img.info):
#                     background = Image.new('RGB', img.size, (255, 255, 255))
#                     background.paste(img, mask=img.split()[3] if img.mode == 'RGBA' else None)
#                     img = background
                
#                 # Redimensionner si l'image est trop grande
#                 max_size = (1920, 1920)
#                 if img.width > max_size[0] or img.height > max_size[1]:
#                     img.thumbnail(max_size, Image.LANCZOS)
                
#                 # Convertir l'image traitée en base64
#                 buffer = io.BytesIO()
#                 img.save(buffer, format='JPEG', quality=85)
#                 img_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
                
#                 _logger.info(f"Image valide téléchargée et traitée: {image_url}")
#                 return img_base64
                
#             except Exception as e:
#                 _logger.error(f"Contenu non valide comme image: {str(e)}")
#                 return False
                
#         except requests.exceptions.RequestException as e:
#             _logger.error(f"Erreur lors de la récupération de l'image {image_url}: {str(e)}")
#             return False
#         except Exception as e:
#             _logger.error(f"Erreur inattendue lors du traitement de l'image {image_url}: {str(e)}")
#             return False


#     @http.route('/api/read_excel_file', methods=['GET'], type='http', auth='none', cors="*")
#     def read_excel_file(self, **kwargs):
#         file_name = kwargs.get('file_name')
#         if not file_name:
#             return Response(
#                 status=400,
#                 content_type='application/json; charset=utf-8',
#                 headers=[('Cache-Control', 'no-store'), ('Pragma', 'no-cache')],
#                 response=json.dumps({"error": "Veuillez fournir le nom du fichier via le paramètre 'file_name'."})
#             )

#         data_dir = os.path.join(os.path.dirname(__file__), '../data')
#         file_path = os.path.join(data_dir, file_name)

#         if not os.path.exists(file_path):
#             return Response(
#                 status=404,
#                 content_type='application/json; charset=utf-8',
#                 headers=[('Cache-Control', 'no-store'), ('Pragma', 'no-cache')],
#                 response=json.dumps({"error": "Fichier non trouvé."})
#             )

#         try:
#             # Charger le fichier Excel
#             workbook = load_workbook(filename=file_path)
#             result = {}

#             for sheet_name in workbook.sheetnames:
#                 sheet = workbook[sheet_name]
#                 sheet_data = []

#                 # Lire les données de chaque feuille
#                 for row in sheet.iter_rows(min_row=2, values_only=True):
#                     if len(row) >= 3:  # Assurez-vous que la ligne a au moins 3 colonnes
#                         code = row[0] # elimine les tirets
#                         code = code.replace('-', '')
#                         nom = row[1]
#                         prix = row[2]

#                         # Formater le prix
#                         if isinstance(prix, str):
#                             prix = prix.replace(' F', '').replace(' ', '').strip()
#                             try:
#                                 prix = int(prix)
#                             except ValueError:
#                                 prix = 0
#                         else:
#                             prix = int(prix) if prix is not None else 0

#                         # Vérifier si le produit existe déjà ( si le code est dans le nom du produit )
#                         product_exists = request.env['product.template'].sudo().search([('name', 'ilike', f"%{code}%")], limit=1)
                        
#                         existe = product_exists and product_exists.id or False

#                         sheet_data.append({
#                             "code": code,
#                             "nom": nom,
#                             "prix": prix,
#                             "existe": bool(existe)
#                         })

#                 result[sheet_name] = sheet_data

#             return Response(
#                 status=200,
#                 content_type='application/json; charset=utf-8',
#                 headers=[('Cache-Control', 'no-store'), ('Pragma', 'no-cache')],
#                 response=json.dumps(result, ensure_ascii=False, indent=4)
#             )

#         except Exception as e:
#             _logger.error(f"Erreur lors du traitement du fichier: {str(e)}", exc_info=True)
#             return Response(
#                 status=500,
#                 content_type='application/json; charset=utf-8',
#                 headers=[('Cache-Control', 'no-store'), ('Pragma', 'no-cache')],
#                 response=json.dumps({"error": f"Erreur lors de la lecture du fichier : {str(e)}"})
#             )



#     def search_web(self, query):
#         try:
#             # Utilisez une API de recherche web pour rechercher le nom du produit
#             # Remplacez 'YOUR_API_KEY' et 'YOUR_CUSTOM_SEARCH_ENGINE_ID' par vos propres clés API
#             api_key = 'AIzaSyC-ffzKUxpc6UD_geVvoTIfM8qaPTlEOg0'
#             custom_search_engine_id = 'AIzaSyC-ffzKUxpc6UD_geVvoTIfM8qaPTlEOg0'
#             url = f'https://www.googleapis.com/customsearch/v1?key={api_key}&cx={custom_search_engine_id}&q={query}'

#             response = requests.get(url)
#             response.raise_for_status()
#             search_results = response.json()

#             if 'items' in search_results and len(search_results['items']) > 0:
#                 return search_results['items'][0]['link']
#             else:
#                 return None
#         except Exception as e:
#             _logger.error(f"Erreur lors de la recherche web: {str(e)}", exc_info=True)
#             return None
        


from odoo import http
from odoo.http import request, Response
import json
import requests
from bs4 import BeautifulSoup
import werkzeug.wrappers
import os
import datetime
import urllib.parse
from pathlib import Path
import base64
import logging
import io
from PIL import Image
from openpyxl import load_workbook

_logger = logging.getLogger(__name__)

class ScraperController(http.Controller):
    
    @http.route('/api/scrape_products_full', methods=['GET'], type='http', auth='none', cors="*")
    def scrape_products_full(self, **kwargs):
        target_url = kwargs.get('target_url')
        fetch_details = kwargs.get('fetch_details', 'false').lower() == 'true'
        save_to_file = kwargs.get('save_to_file', 'true').lower() == 'true'
        
        if not target_url:
            return werkzeug.wrappers.Response(
                status=400,
                content_type='application/json; charset=utf-8',
                headers=[('Cache-Control', 'no-store'), ('Pragma', 'no-cache')],
                response=json.dumps({"error": "Veuillez fournir l'URL cible via le paramètre 'target_url'."})
            )

        try:
            response = requests.get(target_url, timeout=10)
            response.raise_for_status()
            soup = BeautifulSoup(response.text, 'html.parser')
            products = soup.select("ul.products li.product")
            product_data = []

            for product in products:
                title = product.select_one("h2.woocommerce-loop-product__title")
                link = product.select_one("a.woocommerce-LoopProduct-link")
                image = product.select_one("img")
                title_text = title.text.replace('  ', ' ')
                
                product_info = {
                    "title": title_text.strip() if title else None,
                    "link": link["href"] if link else None,
                    "image": image.get("src") if image else None,
                }

                # Détails du produit
                if fetch_details and product_info["link"]:
                    try:
                        details_produits = self.fetch_page_text(target_url=product_info["link"])
                        product_info["summary_text"] = details_produits.get("summary_text", None)
                        product_info["images"] = details_produits.get("images", [])
                        product_info["categories"] = details_produits.get("categories", [])
                        product_info["tags"] = details_produits.get("tags", [])
                    except Exception as e:
                        product_info["detail_fetch_error"] = str(e)

                product_data.append(product_info)

            page_data = {
                "url": target_url,
                "scrape_date": datetime.datetime.now().isoformat(),
                "produits_nbre": len(product_data),
                "produits": product_data
            }

            # Vérifier l'existence des produits dans la base de données
            product_titles = [p['title'] for p in product_data]
            existing_products = request.env['product.product'].sudo().search([('name', 'in', product_titles)])
            existing_product_titles = set(existing_products.mapped('name'))

            for product in product_data:
                product['exists_in_db'] = product['title'] in existing_product_titles

            file_path = None
            if save_to_file:
                data_dir = os.path.join(os.path.dirname(__file__), '../data')
                os.makedirs(data_dir, exist_ok=True)
                domain = urllib.parse.urlparse(target_url).netloc.replace('.', '_')
                parsed_url = urllib.parse.urlparse(target_url)
                category = parsed_url.path.split('/')[-2] if parsed_url.path.count('/') > 1 else 'uncategorized'
                
                domain = parsed_url.netloc.replace('.', '_')
                timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"products_{domain}_{category}.json"
                file_path = os.path.join(data_dir, filename)
                
                with open(file_path, 'w', encoding='utf-8') as f:
                    json.dump(page_data, f, ensure_ascii=False, indent=4)
                
                page_data["file_saved"] = file_path

            return werkzeug.wrappers.Response(
                status=200,
                content_type='application/json; charset=utf-8',
                headers=[('Cache-Control', 'no-store'), ('Pragma', 'no-cache')],
                response=json.dumps(page_data, ensure_ascii=False, indent=4)
            )

        except requests.exceptions.RequestException as e:
            return werkzeug.wrappers.Response(
                status=500,
                content_type='application/json; charset=utf-8',
                headers=[('Cache-Control', 'no-store'), ('Pragma', 'no-cache')],
                response=json.dumps({"error": f"Erreur lors de la requête : {str(e)}"})
            )
        except Exception as e:
            return werkzeug.wrappers.Response(
                status=500,
                content_type='application/json; charset=utf-8',
                headers=[('Cache-Control', 'no-store'), ('Pragma', 'no-cache')],
                response=json.dumps({"error": f"Erreur inconnue : {str(e)}"})
            )

    def fetch_page_text(self, target_url):
        """Fetches the text from a specific section of a webpage."""
        if not target_url:
            return {
                "error": "Veuillez fournir l'URL cible via le paramètre 'target_url'."
            }

        try:
            response = requests.get(target_url, timeout=10)
            response.raise_for_status()
            soup = BeautifulSoup(response.text, 'html.parser')

            # Extraire le texte de la section spécifique
            summary_section = soup.select_one('div.summary.entry-summary')
            if not summary_section:
                return {
                    "error": "La section spécifiée n'a pas été trouvée."
                }

            # Extraire le titre
            title_element = summary_section.select_one('h1.product_title.entry-title')
            title_text = title_element.get_text(strip=True) if title_element else "Titre non trouvé"

            # Extraire la description courte
            description_section = summary_section.select_one('div.woocommerce-product-details__short-description')
            if description_section:
                paragraphs = description_section.find_all('p')
                summary_text = "\n".join(p.get_text(strip=True) for p in paragraphs)
            else:
                summary_text = summary_section.get_text(separator='\n', strip=True)

            # Extraire les images de la galerie
            gallery_section = soup.select_one('div.woocommerce-product-gallery')
            images = []
            if gallery_section:
                for img in gallery_section.select('div.woocommerce-product-gallery__image img'):
                    images.append({
                        "src": img.get("data-large_image") or img.get("src")
                    })

            # Extraire la catégorie et les tags
            category_element = summary_section.select_one('span.posted_in')
            categories = [a.get_text(strip=True) for a in category_element.select('a')] if category_element else []

            tag_element = summary_section.select_one('span.tagged_as')
            tags = [a.get_text(strip=True) for a in tag_element.select('a')] if tag_element else []

            page_data = {
                "title": title_text,
                "summary_text": summary_text,
                "images": images,
                "categories": categories,
                "tags": tags
            }

            return page_data

        except requests.exceptions.RequestException as e:
            return {
                "error": f"Erreur lors de la requête : {str(e)}"
            }
        except Exception as e:
            return {
                "error": f"Erreur inconnue : {str(e)}"
            }

    @http.route('/api/read_product_files', methods=['GET'], type='http', auth='none', cors="*")
    def read_product_files(self, **kwargs):
        file_name = kwargs.get('file_name')
        if not file_name:
            return werkzeug.wrappers.Response(
                status=400,
                content_type='application/json; charset=utf-8',
                headers=[('Cache-Control', 'no-store'), ('Pragma', 'no-cache')],
                response=json.dumps({"error": "Veuillez fournir le nom du fichier via le paramètre 'file_name'."})
            )

        data_dir = os.path.join(os.path.dirname(__file__), '../data')
        file_path = os.path.join(data_dir, file_name)

        if not os.path.exists(file_path):
            return werkzeug.wrappers.Response(
                status=404,
                content_type='application/json; charset=utf-8',
                headers=[('Cache-Control', 'no-store'), ('Pragma', 'no-cache')],
                response=json.dumps({"error": "Fichier non trouvé."})
            )

        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                file_data = json.load(f)

            return werkzeug.wrappers.Response(
                status=200,
                content_type='application/json; charset=utf-8',
                headers=[('Cache-Control', 'no-store'), ('Pragma', 'no-cache')],
                response=json.dumps(file_data['produits'], ensure_ascii=False, indent=4)
            )
        except Exception as e:
            return werkzeug.wrappers.Response(
                status=500,
                content_type='application/json; charset=utf-8',
                headers=[('Cache-Control', 'no-store'), ('Pragma', 'no-cache')],
                response=json.dumps({"error": f"Erreur lors de la lecture du fichier : {str(e)}"})
            )

    def find_existing_product(self, product_data):
        """
        Recherche un produit existant par nom, référence ou code modèle.
        Retourne le produit trouvé ou None.
        """
        product_name = product_data.get('title', '').strip()
        
        if not product_name:
            return None
        
        # 1. Recherche par nom exact
        existing_product = request.env['product.template'].sudo().search([
            ('name', '=', product_name)
        ], limit=1)
        
        if existing_product:
            _logger.info(f"Produit trouvé par nom exact: {existing_product.name} (ID: {existing_product.id})")
            return existing_product
        
        # 2. Recherche par nom similaire (ilike)
        existing_product = request.env['product.template'].sudo().search([
            ('name', 'ilike', product_name)
        ], limit=1)
        
        if existing_product:
            _logger.info(f"Produit trouvé par nom similaire: {existing_product.name} (ID: {existing_product.id})")
            return existing_product
        
        # 3. Recherche par référence (default_code) si disponible
        if product_data.get('code') or product_data.get('reference'):
            product_code = product_data.get('code') or product_data.get('reference')
            existing_product = request.env['product.template'].sudo().search([
                ('default_code', '=', product_code)
            ], limit=1)
            
            if existing_product:
                _logger.info(f"Produit trouvé par référence: {existing_product.default_code} (ID: {existing_product.id})")
                return existing_product
        
        # 4. Recherche par code modèle extrait du nom
        try:
            # Utiliser la logique d'extraction de code modèle du modèle product.template
            product_template_model = request.env['product.template'].sudo()
            if hasattr(product_template_model, '_extract_model_code'):
                extracted_code = product_template_model._extract_model_code(product_name)
                if extracted_code:
                    existing_product = request.env['product.template'].sudo().search([
                        '|',
                        ('default_code', 'ilike', f'%{extracted_code}%'),
                        ('name', 'ilike', f'%{extracted_code}%')
                    ], limit=1)
                    
                    if existing_product:
                        _logger.info(f"Produit trouvé par code modèle extrait: {extracted_code} -> {existing_product.name} (ID: {existing_product.id})")
                        return existing_product
        except Exception as e:
            _logger.warning(f"Erreur lors de l'extraction du code modèle: {str(e)}")
        
        return None

    def update_existing_product(self, existing_product, product_data):
        """
        Met à jour un produit existant avec les nouvelles données.
        """
        try:
            update_vals = {}
            
            # Mettre à jour la description si elle est fournie
            if product_data.get('summary_text'):
                update_vals['description'] = product_data['summary_text']
            
            # Mettre à jour les prix si fournis
            if product_data.get('price'):
                try:
                    update_vals['list_price'] = float(product_data['price'])
                except (ValueError, TypeError):
                    pass
            
            if product_data.get('cost'):
                try:
                    update_vals['standard_price'] = float(product_data['cost'])
                except (ValueError, TypeError):
                    pass
            
            # Mettre à jour les champs booléens
            for field in ['en_promo', 'is_preorder', 'is_creditorder']:
                if field in product_data:
                    update_vals[field] = bool(product_data[field])
            
            # Mettre à jour les prix spéciaux
            for price_field in ['preorder_price', 'promo_price', 'creditorder_price']:
                if product_data.get(price_field):
                    try:
                        update_vals[price_field] = float(product_data[price_field])
                    except (ValueError, TypeError):
                        pass
            
            # Mettre à jour le rang
            if product_data.get('rang'):
                try:
                    update_vals['rang'] = int(product_data['rang'])
                except (ValueError, TypeError):
                    pass
            
            # Mettre à jour la catégorie si fournie
            if product_data.get('categories'):
                categ_id = self.get_category_id(product_data['categories'])
                if categ_id:
                    update_vals['categ_id'] = categ_id
            
            # Mettre à jour les tags si fournis
            if product_data.get('tags'):
                tag_ids = self.get_tag_ids(product_data['tags'])
                if tag_ids:
                    update_vals['product_tag_ids'] = [(6, 0, tag_ids)]
            
            # Mettre à jour l'image principale si fournie
            if product_data.get('image'):
                image_base64 = self.get_image_base64(product_data['image'])
                if image_base64:
                    update_vals['image_1920'] = image_base64
            
            # Mettre à jour les images supplémentaires
            if product_data.get('images'):
                images = product_data['images']
                image_base64_list = []
                
                for img in images:
                    if img.get('src'):
                        img_base64 = self.get_image_base64(img['src'])
                        if img_base64:
                            image_base64_list.append(img_base64)
                
                # Mettre à jour les champs d'images supplémentaires
                if image_base64_list:
                    if len(image_base64_list) > 0:
                        update_vals['image_1'] = image_base64_list[0]
                    if len(image_base64_list) > 1:
                        update_vals['image_2'] = image_base64_list[1]
                    if len(image_base64_list) > 2:
                        update_vals['image_3'] = image_base64_list[2]
                    if len(image_base64_list) > 3:
                        update_vals['image_4'] = image_base64_list[3]
            
            # Effectuer la mise à jour si des valeurs ont été préparées
            if update_vals:
                existing_product.write(update_vals)
                _logger.info(f"Produit mis à jour: {existing_product.name} (ID: {existing_product.id}) avec {len(update_vals)} champs")
                return True
            else:
                _logger.info(f"Aucune mise à jour nécessaire pour: {existing_product.name}")
                return False
                
        except Exception as e:
            _logger.error(f"Erreur lors de la mise à jour du produit {existing_product.name}: {str(e)}", exc_info=True)
            return False

    @http.route('/api/read_product_file', methods=['GET'], type='http', auth='none', cors="*")
    def read_product_file(self, **kwargs):
        file_name = kwargs.get('file_name')
        if not file_name:
            return werkzeug.wrappers.Response(
                status=400,
                content_type='application/json; charset=utf-8',
                headers=[('Cache-Control', 'no-store'), ('Pragma', 'no-cache')],
                response=json.dumps({"error": "Veuillez fournir le nom du fichier via le paramètre 'file_name'."})
            )

        data_dir = os.path.join(os.path.dirname(__file__), '../data')
        file_path = os.path.join(data_dir, file_name)

        if not os.path.exists(file_path):
            return werkzeug.wrappers.Response(
                status=404,
                content_type='application/json; charset=utf-8',
                headers=[('Cache-Control', 'no-store'), ('Pragma', 'no-cache')],
                response=json.dumps({"error": "Fichier non trouvé."})
            )

        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                file_data = json.load(f)

            produits = file_data.get('produits', [])
            created_products = []
            updated_products = []
            skipped_products = []

            # Obtenir la catégorie par défaut avant de commencer
            default_category_id = self.ensure_default_category()
            _logger.info(f"Catégorie par défaut ID: {default_category_id}")

            # Utiliser un savepoint pour pouvoir annuler en cas d'erreur
            with request.env.cr.savepoint():
                for produit in produits:
                    if not produit.get('title'):
                        skipped_products.append("Produit sans titre")
                        continue
                    
                    _logger.info(f"Traitement du produit: {produit['title']}")

                    try:
                        # Rechercher un produit existant
                        existing_product = self.find_existing_product(produit)
                        
                        if existing_product:
                            # Mettre à jour le produit existant
                            _logger.info(f"Produit existant trouvé: {existing_product.name} (ID: {existing_product.id})")
                            
                            update_success = self.update_existing_product(existing_product, produit)
                            if update_success:
                                updated_products.append(f"{produit['title']} -> {existing_product.name}")
                            else:
                                updated_products.append(f"{produit['title']} (aucune modification)")
                        else:
                            # Créer un nouveau produit
                            _logger.info(f"Création d'un nouveau produit: {produit['title']}")
                            
                            # Vérifier et récupérer l'image
                            image_base64 = self.get_image_base64(produit.get('image'))
                            
                            # Récupérer les images supplémentaires
                            images = produit.get('images', [])
                            image_base64_list = []
                            
                            # Traiter chaque image avec gestion des erreurs
                            for img in images:
                                if img.get('src'):
                                    img_base64 = self.get_image_base64(img['src'])
                                    if img_base64:
                                        image_base64_list.append(img_base64)
                            
                            # S'assurer que categ_id est toujours défini
                            categ_id = self.get_category_id(produit.get('categories', []))
                            if not categ_id:
                                categ_id = default_category_id
                            
                            _logger.info(f"Catégorie utilisée pour le produit: {categ_id}")
                            
                            # Création d'un nouveau modèle de produit
                            product_vals = {
                                'name': produit['title'],
                                'sale_ok': False,
                                'purchase_ok': False,
                                'detailed_type': 'product',
                                'description': produit.get('summary_text'),
                                'list_price': float(produit.get('price', 0.0)),
                                'standard_price': float(produit.get('cost', 0.0)),
                                'en_promo': bool(produit.get('en_promo', False)),
                                'is_preorder': bool(produit.get('is_preorder', False)),
                                'preorder_price': float(produit.get('preorder_price', 0.0)),
                                'promo_price': float(produit.get('promo_price', 0.0)),
                                'is_creditorder': bool(produit.get('is_creditorder', False)),
                                'creditorder_price': float(produit.get('creditorder_price', 0.0)),
                                'rang': int(produit.get('rang', 0)),
                                'categ_id': categ_id,
                            }
                            
                            # Ajouter les tags si disponibles
                            tag_ids = self.get_tag_ids(produit.get('tags', []))
                            if tag_ids:
                                product_vals['product_tag_ids'] = [(6, 0, tag_ids)]
                            
                            # Ajouter les images supplémentaires si disponibles
                            if image_base64_list:
                                if len(image_base64_list) > 0:
                                    product_vals['image_1'] = image_base64_list[0]
                                if len(image_base64_list) > 1:
                                    product_vals['image_2'] = image_base64_list[1]
                                if len(image_base64_list) > 2:
                                    product_vals['image_3'] = image_base64_list[2]
                                if len(image_base64_list) > 3:
                                    product_vals['image_4'] = image_base64_list[3]

                            # N'ajouter l'image principale que si elle est valide
                            if image_base64:
                                product_vals['image_1920'] = image_base64
                            
                            _logger.info(f"Création du produit avec les valeurs: {product_vals.keys()}")
                            _logger.info(f"Catégorie ID utilisée: {product_vals.get('categ_id')}")
                            
                            product_template = request.env['product.template'].sudo().create(product_vals)
                            _logger.info(f"Nouveau produit créé avec ID: {product_template.id}")
                            created_products.append(produit['title'])

                    except Exception as e:
                        _logger.error(f"Erreur lors du traitement du produit {produit['title']}: {str(e)}", exc_info=True)
                        skipped_products.append(f"{produit['title']} (Erreur: {str(e)})")

            result = {
                "success": True,
                "created": len(created_products),
                "updated": len(updated_products),
                "skipped": len(skipped_products),
                "created_products": created_products,
                "updated_products": updated_products,
                "skipped_products": skipped_products
            }

            return werkzeug.wrappers.Response(
                status=200,
                content_type='application/json; charset=utf-8',
                headers=[('Cache-Control', 'no-store'), ('Pragma', 'no-cache')],
                response=json.dumps(result, ensure_ascii=False, indent=4)
            )

        except Exception as e:
            _logger.error(f"Erreur lors du traitement du fichier: {str(e)}", exc_info=True)
            return werkzeug.wrappers.Response(
                status=500,
                content_type='application/json; charset=utf-8',
                headers=[('Cache-Control', 'no-store'), ('Pragma', 'no-cache')],
                response=json.dumps({"error": f"Erreur lors de la lecture du fichier : {str(e)}"})
            )

    def ensure_default_category(self):
        """
        S'assure qu'une catégorie par défaut existe et retourne son ID.
        Cette méthode ne devrait jamais retourner False ou None.
        """
        # Catégorie par défaut à utiliser
        default_category_name = "Non catégorisé"
        
        # Rechercher la catégorie par défaut
        default_category = request.env['product.category'].sudo().search([('name', '=', default_category_name)], limit=1)
        
        if not default_category:
            # Créer la catégorie par défaut si elle n'existe pas
            _logger.info(f"Création de la catégorie par défaut: {default_category_name}")
            try:
                default_category = request.env['product.category'].sudo().create({'name': default_category_name})
                _logger.info(f"Catégorie par défaut créée avec ID: {default_category.id}")
            except Exception as e:
                _logger.error(f"Erreur lors de la création de la catégorie par défaut: {str(e)}")
                # En cas d'échec, utiliser la catégorie "All" (ID 1) qui existe toujours dans Odoo
                return 1
        
        _logger.info(f"Catégorie par défaut trouvée avec ID: {default_category.id}")
        return default_category.id

    def get_category_id(self, categories):
        """
        Récupère ou crée une catégorie de produit.
        Retourne un ID de catégorie valide ou None si une erreur se produit.
        """
        try:
            if not categories or not categories[0]:
                _logger.warning("Aucune catégorie fournie")
                return None
            
            category_name = categories[0]  # Prendre la première catégorie
            _logger.info(f"Recherche de la catégorie: {category_name}")
            
            category = request.env['product.category'].sudo().search([('name', '=', category_name)], limit=1)
            
            if not category:
                # Créer la catégorie si elle n'existe pas
                _logger.info(f"Création de la catégorie: {category_name}")
                category = request.env['product.category'].sudo().create({'name': category_name})
            
            _logger.info(f"Catégorie trouvée/créée avec ID: {category.id}")
            return category.id
        except Exception as e:
            _logger.error(f"Erreur lors de la récupération/création de la catégorie: {str(e)}")
            return None

    def get_tag_ids(self, tags):
        """
        Récupère ou crée des tags de produit.
        Retourne une liste d'IDs de tags.
        """
        tag_ids = []
        if not tags:
            return tag_ids
        
        for tag_name in tags:
            try:
                tag = request.env['product.tag'].sudo().search([('name', '=', tag_name)], limit=1)
                if tag:
                    tag_ids.append(tag.id)
                else:
                    new_tag = request.env['product.tag'].sudo().create({'name': tag_name})
                    tag_ids.append(new_tag.id)
            except Exception as e:
                _logger.error(f"Erreur lors de la récupération/création du tag {tag_name}: {str(e)}")
        
        return tag_ids

    def get_image_base64(self, image_url):
        """Récupère une image depuis une URL et la convertit en base64 avec vérification."""
        if not image_url:
            _logger.warning("Aucune URL d'image fournie")
            return False
        
        try:
            _logger.info(f"Téléchargement de l'image depuis: {image_url}")
            response = requests.get(image_url, timeout=10)
            response.raise_for_status()
            
            # Vérifier si le contenu est une image valide
            content = response.content
            try:
                # Essayer d'ouvrir l'image avec PIL pour vérifier qu'elle est valide
                img = Image.open(io.BytesIO(content))
                img.verify()  # Vérifier que l'image est valide
                
                # Réouvrir l'image pour la traiter (après verify(), l'image est fermée)
                img = Image.open(io.BytesIO(content))
                
                # Convertir en RGB si nécessaire (pour les images avec canal alpha)
                if img.mode in ('RGBA', 'LA') or (img.mode == 'P' and 'transparency' in img.info):
                    background = Image.new('RGB', img.size, (255, 255, 255))
                    background.paste(img, mask=img.split()[3] if img.mode == 'RGBA' else None)
                    img = background
                
                # Redimensionner si l'image est trop grande
                max_size = (1920, 1920)
                if img.width > max_size[0] or img.height > max_size[1]:
                    img.thumbnail(max_size, Image.LANCZOS)
                
                # Convertir l'image traitée en base64
                buffer = io.BytesIO()
                img.save(buffer, format='JPEG', quality=85)
                img_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
                
                _logger.info(f"Image valide téléchargée et traitée: {image_url}")
                return img_base64
                
            except Exception as e:
                _logger.error(f"Contenu non valide comme image: {str(e)}")
                return False
            
        except requests.exceptions.RequestException as e:
            _logger.error(f"Erreur lors de la récupération de l'image {image_url}: {str(e)}")
            return False
        except Exception as e:
            _logger.error(f"Erreur inattendue lors du traitement de l'image {image_url}: {str(e)}")
            return False

    @http.route('/api/read_excel_file', methods=['GET'], type='http', auth='none', cors="*")
    def read_excel_file(self, **kwargs):
        file_name = kwargs.get('file_name')
        if not file_name:
            return Response(
                status=400,
                content_type='application/json; charset=utf-8',
                headers=[('Cache-Control', 'no-store'), ('Pragma', 'no-cache')],
                response=json.dumps({"error": "Veuillez fournir le nom du fichier via le paramètre 'file_name'."})
            )

        data_dir = os.path.join(os.path.dirname(__file__), '../data')
        file_path = os.path.join(data_dir, file_name)

        if not os.path.exists(file_path):
            return Response(
                status=404,
                content_type='application/json; charset=utf-8',
                headers=[('Cache-Control', 'no-store'), ('Pragma', 'no-cache')],
                response=json.dumps({"error": "Fichier non trouvé."})
            )

        try:
            # Charger le fichier Excel
            workbook = load_workbook(filename=file_path)
            result = {}

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_data = []

                # Lire les données de chaque feuille
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if len(row) >= 3:  # Assurez-vous que la ligne a au moins 3 colonnes
                        code = row[0]  # elimine les tirets
                        code = code.replace('-', '')
                        nom = row[1]
                        prix = row[2]

                        # Formater le prix
                        if isinstance(prix, str):
                            prix = prix.replace(' F', '').replace(' ', '').strip()
                            try:
                                prix = int(prix)
                            except ValueError:
                                prix = 0
                        else:
                            prix = int(prix) if prix is not None else 0

                        # Vérifier si le produit existe déjà (si le code est dans le nom du produit)
                        product_exists = request.env['product.template'].sudo().search([('name', 'ilike', f"%{code}%")], limit=1)
                        
                        existe = product_exists and product_exists.id or False

                        sheet_data.append({
                            "code": code,
                            "nom": nom,
                            "prix": prix,
                            "existe": bool(existe)
                        })

                result[sheet_name] = sheet_data

            return Response(
                status=200,
                content_type='application/json; charset=utf-8',
                headers=[('Cache-Control', 'no-store'), ('Pragma', 'no-cache')],
                response=json.dumps(result, ensure_ascii=False, indent=4)
            )

        except Exception as e:
            _logger.error(f"Erreur lors du traitement du fichier: {str(e)}", exc_info=True)
            return Response(
                status=500,
                content_type='application/json; charset=utf-8',
                headers=[('Cache-Control', 'no-store'), ('Pragma', 'no-cache')],
                response=json.dumps({"error": f"Erreur lors de la lecture du fichier : {str(e)}"})
            )

    def search_web(self, query):
        try:
            # Utilisez une API de recherche web pour rechercher le nom du produit
            # Remplacez 'YOUR_API_KEY' et 'YOUR_CUSTOM_SEARCH_ENGINE_ID' par vos propres clés API
            api_key = 'AIzaSyC-ffzKUxpc6UD_geVvoTIfM8qaPTlEOg0'
            custom_search_engine_id = 'AIzaSyC-ffzKUxpc6UD_geVvoTIfM8qaPTlEOg0'
            url = f'https://www.googleapis.com/customsearch/v1?key={api_key}&cx={custom_search_engine_id}&q={query}'
            response = requests.get(url)
            response.raise_for_status()
            search_results = response.json()

            if 'items' in search_results and len(search_results['items']) > 0:
                return search_results['items'][0]['link']
            else:
                return None
        except Exception as e:
            _logger.error(f"Erreur lors de la recherche web: {str(e)}", exc_info=True)
            return None
